#!/usr/bin/env python3
"""Script to check Excel template structure"""

import openpyxl

# Load the Excel template
wb = openpyxl.load_workbook('SA - Empty.xlsx')

# Print all sheet names
print("Sheet names:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# For each sheet, print its structure
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")

    # Print first 30 rows with content
    print("\nContent (first 30 rows):")
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=False), 1):
        row_values = [cell.value for cell in row if cell.value is not None]
        if row_values:
            # Also show cell coordinates
            cells_info = [(cell.coordinate, cell.value) for cell in row if cell.value is not None]
            print(f"Row {row_idx}: {cells_info}")

wb.close()
